import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_student_excel_template():
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define colors for styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    required_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    optional_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    
    # Create Student Data sheet
    ws_student = wb.create_sheet("Student_Data")
    
    # Student data headers
    student_headers = [
        "Student_First_Name*", "Student_Middle_Name", "Student_Last_Name*", "Student_Display_Name",
        "Student_Username*", "Student_Phone*", "Student_Gender*", "Student_Date_of_Birth*",
        "Student_Tazkira_No", "Student_Bio", "Admission_Date*", "Class_ID*",
        "Blood_Group", "Nationality", "Religion", "Caste", "Aadhar_No",
        "Student_Tazkira_No_2", "Bank_Account_No", "Bank_Name", "IFSC_Code", "Previous_School"
    ]
    
    # Add headers to student sheet
    for col, header in enumerate(student_headers, 1):
        cell = ws_student.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        if "*" in header:
            cell.fill = required_fill
        else:
            cell.fill = optional_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create Parent Data sheet
    ws_parent = wb.create_sheet("Parent_Data")
    
    # Parent data headers
    parent_headers = [
        "Parent_First_Name*", "Parent_Middle_Name", "Parent_Last_Name*", "Parent_Display_Name",
        "Parent_Username*", "Parent_Phone*", "Parent_Gender*", "Parent_Birth_Date*",
        "Parent_Tazkira_No", "Parent_Bio", "Occupation", "Annual_Income",
        "Education", "Employer", "Designation", "Work_Phone", "Emergency_Contact",
        "Relationship", "Is_Guardian", "Is_Emergency_Contact"
    ]
    
    # Add headers to parent sheet
    for col, header in enumerate(parent_headers, 1):
        cell = ws_parent.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        if "*" in header:
            cell.fill = required_fill
        else:
            cell.fill = optional_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create Address Data sheet
    ws_address = wb.create_sheet("Address_Data")
    
    # Address data headers
    address_headers = [
        "Origin_Address", "Origin_City", "Origin_State", "Origin_Province",
        "Origin_Country", "Origin_Postal_Code", "Current_Address", "Current_City",
        "Current_State", "Current_Province", "Current_Country", "Current_Postal_Code"
    ]
    
    # Add headers to address sheet
    for col, header in enumerate(address_headers, 1):
        cell = ws_address.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = optional_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["STUDENT DATA INSERTION TEMPLATE - INSTRUCTIONS"],
        [""],
        ["REQUIRED FIELDS (marked with *):"],
        ["• Student_First_Name, Student_Last_Name, Student_Username, Student_Phone, Student_Gender, Student_Date_of_Birth"],
        ["• Parent_First_Name, Parent_Last_Name, Parent_Username, Parent_Phone, Parent_Gender, Parent_Birth_Date"],
        ["• Admission_Date, Class_ID"],
        [""],
        ["DATA FORMATS:"],
        ["• Dates: YYYY-MM-DD format (e.g., 2024-01-15)"],
        ["• Phone: Include country code (e.g., +923001234567)"],
        ["• Gender: MALE, FEMALE, or OTHER"],
        ["• Class_ID: Numeric value (e.g., 1, 2, 3)"],
        ["• Relationship: Father, Mother, Guardian, or Other"],
        ["• Boolean fields: true or false"],
        [""],
        ["SAMPLE DATA ROW:"],
        ["Student_First_Name: Ahmed"],
        ["Student_Last_Name: Khan"],
        ["Student_Username: ahmed_khan_2024"],
        ["Student_Phone: +923001234567"],
        ["Student_Gender: MALE"],
        ["Student_Date_of_Birth: 2010-05-15"],
        ["Parent_First_Name: Muhammad"],
        ["Parent_Last_Name: Khan"],
        ["Parent_Username: muhammad_khan_parent"],
        ["Parent_Phone: +923001234568"],
        ["Parent_Gender: MALE"],
        ["Parent_Birth_Date: 1980-03-20"],
        ["Admission_Date: 2024-01-15"],
        ["Class_ID: 1"],
        [""],
        ["IMPORTANT NOTES:"],
        ["• Each row represents one student with their parent data"],
        ["• Keep data consistent across all sheets for the same student"],
        ["• Usernames must be unique"],
        ["• Phone numbers should be in international format"],
        ["• All required fields must be filled before submission"],
        [""],
        ["VALIDATION RULES:"],
        ["• Username: Alphanumeric, no spaces"],
        ["• Phone: Must include country code"],
        ["• Dates: Valid date format"],
        ["• Gender: Must be one of the specified values"],
        ["• Class_ID: Must be a valid class ID from your system"],
        [""],
        ["COLOR CODING:"],
        ["• Red headers: Required fields"],
        ["• Green headers: Optional fields"],
        ["• Blue headers: Address fields"]
    ]
    
    # Add instructions to sheet
    for row, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row, column=1, value=instruction[0])
        if row == 1:
            cell.font = Font(bold=True, size=16, color="366092")
        elif "REQUIRED FIELDS" in instruction[0] or "DATA FORMATS" in instruction[0] or "SAMPLE DATA" in instruction[0] or "IMPORTANT NOTES" in instruction[0] or "VALIDATION RULES" in instruction[0] or "COLOR CODING" in instruction[0]:
            cell.font = Font(bold=True, color="366092")
    
    # Create Sample Data sheet
    ws_sample = wb.create_sheet("Sample_Data")
    
    # Sample data
    sample_data = [
        ["Ahmed", "", "Khan", "Ahmed Khan", "ahmed_khan_2024", "+923001234567", "MALE", "2010-05-15", "", "Student bio", "2024-01-15", "1", "A+", "Pakistani", "Islam", "", "", "", "", "", "", ""],
        ["Fatima", "", "Ali", "Fatima Ali", "fatima_ali_2024", "+923001234568", "FEMALE", "2011-08-20", "", "Student bio", "2024-01-15", "2", "B+", "Pakistani", "Islam", "", "", "", "", "", "", ""],
        ["Hassan", "Muhammad", "Raza", "Hassan Raza", "hassan_raza_2024", "+923001234569", "MALE", "2009-12-10", "", "Student bio", "2024-01-15", "3", "O+", "Pakistani", "Islam", "", "", "", "", "", "", ""]
    ]
    
    # Add sample data headers
    for col, header in enumerate(student_headers, 1):
        cell = ws_sample.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        if "*" in header:
            cell.fill = required_fill
        else:
            cell.fill = optional_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws_sample.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths for better readability
    for ws in [ws_student, ws_parent, ws_address, ws_sample]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    filename = "Student_Data_Template.xlsx"
    wb.save(filename)
    print(f"✅ Excel template created successfully: {filename}")
    print(f"📁 File location: {os.path.abspath(filename)}")
    
    return filename

if __name__ == "__main__":
    try:
        filename = create_student_excel_template()
        print("\n📋 Template includes:")
        print("• Student_Data sheet with all student fields")
        print("• Parent_Data sheet with all parent fields") 
        print("• Address_Data sheet with address fields")
        print("• Instructions sheet with usage guidelines")
        print("• Sample_Data sheet with example entries")
        print("\n🎨 Features:")
        print("• Color-coded headers (Red = Required, Green = Optional)")
        print("• Proper column widths for readability")
        print("• Sample data for reference")
        print("• Comprehensive instructions")
    except Exception as e:
        print(f"❌ Error creating template: {e}") 